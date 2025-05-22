
from services.line_bot import download_image_via_http
from services.state import user_data
import httpx
import json
# 儲存資料與下載圖片
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
from services.data_manager import load_data
from config import LOG_DIR, CASE_DIR
from logger import log
from services.state import remove_collecting_user, remove_case


# 指定 Excel 檔案

def check_excel_file():
    EXCEL_PATH = os.path.join(CASE_DIR, "客戶射出穩定度調查.xlsx")
    # 初始化 Excel
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["服務反饋日期", "服務人員", "客戶", "機號", "出廠日期", "產品",
                   "塑料", "其他備註", "SPC狀態", "其他注意提醒", "檔案資料連結"])  # 欄位名稱
        wb.save(EXCEL_PATH)
    return EXCEL_PATH


def save_to_excel(user_id: str, access_token):
    excel_path = check_excel_file()
    # 取得json
    data = load_data()

    wb = load_workbook(excel_path)
    ws = wb.active

    user_case_message = data['new_case'].get(
        user_id, {}).get('case_message', {})
    images = data['new_case'].get(
        user_id, {}).get('images', [])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    date = datetime.now().strftime("%Y%m%d")
    folder_name = f"{date}-{user_case_message.get('客戶', '客戶名稱') or '客戶名稱'}-{user_case_message.get('機號', '機號') or '機號'}"
    folder_path = os.path.join(CASE_DIR, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    ws.append([
        timestamp,
        user_case_message.get("服務人員", ""),
        user_case_message.get("客戶", ""),
        user_case_message.get("機號", ""),
        user_case_message.get("出廠日期", ""),
        user_case_message.get("產品", ""),
        user_case_message.get("塑料", ""),
        user_case_message.get("其他備註", ""),
    ])
    # ws.append，此時資料已經寫進 ws，max_row 會遞增

    # 圖片處理
    failed_images = []
    for index, image_id in enumerate(images, start=1):
        image_path = os.path.join(folder_path, f"{folder_name}_{index}.jpg")
        if download_image_via_http(image_id, access_token, image_path):
            log(f"[成功] 第 {index} 張圖片下載完成：{image_id}")
        else:
            log(f"[失敗] 第 {index} 張圖片下載失敗：{image_id}")
            failed_images.append({
                "index": index,
                "image_id": image_id,
                "path": image_path
            })

    row_index = ws.max_row  # 取得最後一行

    # 先抓表頭欄位名稱（第1列）
    header = [cell.value for cell in ws[1]]
    # 找出指定欄位名稱的 index（column從1開始）
    column_index = header.index("檔案資料連結") + 1  # Excel 欄位從 1 開始，不是 0！
    folder_absolute_path = os.path.abspath(folder_path)
    cell = ws.cell(row=row_index, column=column_index, value="連結")
    cell.hyperlink = folder_absolute_path
    cell.font = Font(color="0000FF", underline="single")  # 超連結樣式


    # 如果有下載失敗的圖片，儲存成 JSON
    if failed_images:
        failed_log_path = os.path.join(LOG_DIR, "download_failed.json")
        with open(failed_log_path, "w", encoding="utf-8") as f:
            json.dump(failed_images, f, ensure_ascii=False, indent=2)
        log(f"[記錄] 有失敗的圖片，已儲存在：{failed_log_path}")

    wb.save(excel_path)
    # 清除使用者資料
    remove_collecting_user(user_id)
    remove_case(user_id)
